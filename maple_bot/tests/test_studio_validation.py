# Studio GT와 solver trace를 비교하는 채점기를 검증한다.
from __future__ import annotations

import json
import unittest

from openpyxl import load_workbook

from core.puzzle.studio_validation import (
    _annotate_diagnostic_frame,
    _representative_failure_keys,
    score_studio_session,
)


def _write_jsonl(path, rows):
    with path.open("w", encoding="utf-8") as fp:
        for row in rows:
            fp.write(json.dumps(row, ensure_ascii=False))
            fp.write("\n")


class StudioValidationTest(unittest.TestCase):
    def test_score_decomposes_failures_across_retained_raw_box_and_absent_stages(self) -> None:
        from tempfile import TemporaryDirectory
        from pathlib import Path

        with TemporaryDirectory(prefix="studio-validation-bottleneck-") as tmp:
            root = Path(tmp)
            gt_path = root / "gt.jsonl"
            trace_path = root / "trace.jsonl"
            _write_jsonl(
                gt_path,
                [
                    {"run_id": "r1", "frame_id": 0, "target_x": 10, "target_y": 10},
                    {"run_id": "r1", "frame_id": 1, "target_x": 20, "target_y": 20},
                    {"run_id": "r1", "frame_id": 2, "target_x": 41, "target_y": 41},
                    {"run_id": "r1", "frame_id": 3, "target_x": 20, "target_y": 20},
                ],
            )
            candidates = (
                {"candidate_id": "retained", "center": [10, 10], "bbox": [5, 5, 15, 15]},
                {"candidate_id": "raw", "center": [20, 20], "bbox": [15, 15, 25, 25]},
                {"candidate_id": "box", "center": [70, 70], "bbox": [40, 40, 100, 100]},
                {"candidate_id": "absent", "center": [80, 80], "bbox": [75, 75, 85, 85]},
            )
            trace_rows = []
            for frame_index, candidate in enumerate(candidates):
                retained_points = [[10, 10]] if frame_index == 0 else [[100, 100]]
                trace_rows.extend(
                    [
                        {
                            "type": "CANDIDATES",
                            "frame_index": frame_index,
                            "payload": {"candidates": [candidate]},
                        },
                        {
                            "type": "TEMPORAL_SELECTOR",
                            "frame_index": frame_index,
                            "payload": {
                                "debug": {"kinematic_wide_beam_points": retained_points},
                            },
                        },
                        {
                            "type": "SOLVER_VISUAL_TRACE",
                            "frame_index": frame_index,
                            "payload": {
                                "selected_x": 120,
                                "selected_y": 120,
                                "mouse_enabled": False,
                                "candidate_count": 1,
                            },
                        },
                    ]
                )
            _write_jsonl(trace_path, trace_rows)

            result = score_studio_session(gt_path, trace_path, root / "report", pass_distance_px=10.0)

            coverage = result.candidate_coverage
            self.assertEqual(coverage.retained_hypothesis_oracle_frames, 1)
            self.assertEqual(coverage.failed_selector_frames, 1)
            self.assertEqual(coverage.failed_hypothesis_generation_frames, 1)
            self.assertEqual(coverage.failed_box_only_frames, 1)
            self.assertEqual(coverage.failed_candidate_absent_frames, 1)
            self.assertEqual(
                coverage.failed_selector_frames
                + coverage.failed_hypothesis_generation_frames
                + coverage.failed_box_only_frames
                + coverage.failed_candidate_absent_frames,
                result.summary.failed_frames,
            )
            report = result.report_path.read_text(encoding="utf-8")
            self.assertIn("failed_selector_frames: 1", report)
            self.assertIn("failed_hypothesis_generation_frames: 1", report)

    def test_score_reports_candidate_center_and_box_oracle_coverage(self) -> None:
        from tempfile import TemporaryDirectory
        from pathlib import Path

        with TemporaryDirectory(prefix="studio-validation-coverage-") as tmp:
            root = Path(tmp)
            gt_path = root / "gt.jsonl"
            trace_path = root / "trace.jsonl"
            _write_jsonl(
                gt_path,
                [
                    {"run_id": "r1", "frame_id": 0, "target_x": 10, "target_y": 10},
                    {"run_id": "r1", "frame_id": 1, "target_x": 41, "target_y": 41},
                    {"run_id": "r1", "frame_id": 2, "target_x": 20, "target_y": 20},
                ],
            )
            trace_rows = []
            candidate_rows = [
                {"candidate_id": "center", "center": [10, 10], "bbox": [5, 5, 15, 15]},
                {"candidate_id": "box", "center": [70, 70], "bbox": [40, 40, 100, 100]},
                {"candidate_id": "absent", "center": [80, 80], "bbox": [75, 75, 85, 85]},
            ]
            for frame_index, candidate in enumerate(candidate_rows):
                trace_rows.extend(
                    [
                        {
                            "type": "CANDIDATES",
                            "frame_index": frame_index,
                            "payload": {"candidates": [candidate]},
                        },
                        {
                            "type": "SOLVER_VISUAL_TRACE",
                            "frame_index": frame_index,
                            "payload": {
                                "selected_x": 120,
                                "selected_y": 120,
                                "mouse_enabled": False,
                                "candidate_count": 1,
                            },
                        },
                    ]
                )
            _write_jsonl(trace_path, trace_rows)

            result = score_studio_session(gt_path, trace_path, root / "report", pass_distance_px=10.0)

            coverage = result.candidate_coverage
            self.assertEqual(coverage.aligned_target_frames, 3)
            self.assertEqual(coverage.center_oracle_frames, 1)
            self.assertEqual(coverage.box_oracle_frames, 2)
            self.assertEqual(coverage.failed_center_recoverable_frames, 1)
            self.assertEqual(coverage.failed_box_only_frames, 1)
            self.assertEqual(coverage.failed_candidate_absent_frames, 1)
            self.assertIn("candidate_coverage", load_workbook(result.xlsx_path).sheetnames)
            report = result.report_path.read_text(encoding="utf-8")
            self.assertIn("center_oracle_frames: 1", report)
            self.assertIn("failed_box_only_frames: 1", report)

    def test_score_studio_session_writes_jsonl_xlsx_and_summary(self) -> None:
        from tempfile import TemporaryDirectory
        from pathlib import Path

        with TemporaryDirectory(prefix="studio-validation-") as tmp:
            root = Path(tmp)
            gt_path = root / "gt.jsonl"
            trace_path = root / "trace.jsonl"
            out_dir = root / "report"
            _write_jsonl(
                gt_path,
                [
                    {"run_id": "r1", "frame_id": 0, "target_x": 10, "target_y": 10},
                    {"run_id": "r1", "frame_id": 1, "target_x": 20, "target_y": 20},
                ],
            )
            _write_jsonl(
                trace_path,
                [
                    {
                        "type": "SOLVER_VISUAL_TRACE",
                        "frame_index": 0,
                        "payload": {
                            "selected_x": 13,
                            "selected_y": 14,
                            "mouse_enabled": False,
                            "candidate_count": 2,
                        },
                    },
                    {
                        "type": "SOLVER_VISUAL_TRACE",
                        "frame_index": 1,
                        "payload": {
                            "selected_x": 80,
                            "selected_y": 80,
                            "mouse_enabled": False,
                            "candidate_count": 1,
                        },
                    },
                ],
            )

            result = score_studio_session(gt_path, trace_path, out_dir, pass_distance_px=10.0)

            self.assertEqual(result.summary.total_frames, 2)
            self.assertEqual(result.summary.passed_frames, 1)
            self.assertEqual(result.summary.failed_frames, 1)
            self.assertEqual(result.summary.aligned_frames, 2)
            self.assertEqual(result.summary.alignment_missing_frames, 0)
            self.assertEqual(result.summary.alignment_rate, 1.0)
            self.assertEqual(result.summary.aligned_pass_rate, 0.5)
            self.assertTrue(result.score_jsonl.exists())
            self.assertTrue(result.xlsx_path.exists())
            self.assertTrue(result.report_path.exists())
            workbook = load_workbook(result.xlsx_path)
            self.assertIn("summary", workbook.sheetnames)
            self.assertIn("runs", workbook.sheetnames)
            self.assertIn("failure_clusters", workbook.sheetnames)
            self.assertIn("frames", workbook.sheetnames)

    def test_score_uses_final_target_selection_over_identity_trace(self) -> None:
        from tempfile import TemporaryDirectory
        from pathlib import Path

        with TemporaryDirectory(prefix="studio-validation-target-") as tmp:
            root = Path(tmp)
            gt_path = root / "gt.jsonl"
            trace_path = root / "trace.jsonl"
            _write_jsonl(gt_path, [{"run_id": "r1", "frame_id": 0, "target_x": 10, "target_y": 10}])
            _write_jsonl(
                trace_path,
                [
                    {
                        "type": "TARGET_SELECTION",
                        "frame_index": 0,
                        "payload": {
                            "point": [10, 10],
                            "source": "temporal",
                            "reason": "selected_family",
                        },
                    },
                    {
                        "type": "SOLVER_VISUAL_TRACE",
                        "frame_index": 0,
                        "payload": {
                            "selected_x": 100,
                            "selected_y": 100,
                            "candidate_id": "wrong_identity",
                            "mouse_enabled": False,
                            "candidate_count": 2,
                        },
                    },
                ],
            )

            result = score_studio_session(gt_path, trace_path, root / "report", pass_distance_px=10.0)

            self.assertEqual(result.summary.passed_frames, 1)
            self.assertEqual(result.frames[0].selected_x, 10.0)
            self.assertEqual(result.frames[0].selected_y, 10.0)
            self.assertEqual(result.frames[0].candidate_id, "wrong_identity")

    def test_score_studio_session_fails_if_mouse_enabled(self) -> None:
        from tempfile import TemporaryDirectory
        from pathlib import Path

        with TemporaryDirectory(prefix="studio-validation-") as tmp:
            root = Path(tmp)
            gt_path = root / "gt.jsonl"
            trace_path = root / "trace.jsonl"
            out_dir = root / "report"
            _write_jsonl(gt_path, [{"run_id": "r1", "frame_id": 0, "target_x": 10, "target_y": 10}])
            _write_jsonl(
                trace_path,
                [
                    {
                        "type": "SOLVER_VISUAL_TRACE",
                        "frame_index": 0,
                        "payload": {
                            "selected_x": 10,
                            "selected_y": 10,
                            "mouse_enabled": True,
                            "candidate_count": 1,
                        },
                    }
                ],
            )

            result = score_studio_session(gt_path, trace_path, out_dir, pass_distance_px=10.0)

            self.assertEqual(result.summary.passed_frames, 0)
            self.assertEqual(result.frames[0].fail_reason, "mouse_enabled")

    def test_score_aligns_gt_and_solver_by_timestamp(self) -> None:
        from tempfile import TemporaryDirectory
        from pathlib import Path

        with TemporaryDirectory(prefix="studio-validation-time-") as tmp:
            root = Path(tmp)
            gt_path = root / "gt.jsonl"
            trace_path = root / "trace.jsonl"
            _write_jsonl(
                gt_path,
                [
                    {
                        "run_id": "r1",
                        "run_index": 0,
                        "frame_id": 0,
                        "timestamp_ms": 1000,
                        "target_x": 10,
                        "target_y": 20,
                    }
                ],
            )
            _write_jsonl(
                trace_path,
                [
                    {
                        "type": "SOLVER_VISUAL_TRACE",
                        "frame_index": 77,
                        "timestamp_ms": 1030,
                        "payload": {
                            "selected_x": 10,
                            "selected_y": 20,
                            "mouse_enabled": False,
                            "candidate_count": 3,
                            "confidence": 0.8,
                        },
                    }
                ],
            )

            result = score_studio_session(
                gt_path,
                trace_path,
                root / "report",
                pass_distance_px=10.0,
                max_alignment_ms=50.0,
            )

            self.assertEqual(77, result.frames[0].solver_frame_index)
            self.assertEqual(30.0, result.frames[0].timestamp_delta_ms)
            self.assertTrue(result.frames[0].passed)

    def test_score_prefers_explicit_lockstep_solver_frame_link(self) -> None:
        from tempfile import TemporaryDirectory
        from pathlib import Path

        with TemporaryDirectory(prefix="studio-validation-lockstep-") as tmp:
            root = Path(tmp)
            gt_path = root / "gt.jsonl"
            trace_path = root / "trace.jsonl"
            _write_jsonl(
                gt_path,
                [
                    {
                        "run_id": "r1",
                        "run_index": 0,
                        "frame_id": 0,
                        "solver_frame_index": 77,
                        "timestamp_ms": 1000,
                        "target_x": 10,
                        "target_y": 20,
                    }
                ],
            )
            _write_jsonl(
                trace_path,
                [
                    {
                        "type": "SOLVER_VISUAL_TRACE",
                        "frame_index": 77,
                        "timestamp_ms": 5000,
                        "payload": {
                            "selected_x": 10,
                            "selected_y": 20,
                            "mouse_enabled": False,
                            "candidate_count": 1,
                        },
                    }
                ],
            )

            result = score_studio_session(
                gt_path,
                trace_path,
                root / "report",
                pass_distance_px=10.0,
                max_alignment_ms=1.0,
            )

            self.assertEqual(result.frames[0].solver_frame_index, 77)
            self.assertEqual(result.summary.alignment_rate, 1.0)
            self.assertTrue(result.frames[0].passed)

    def test_score_reports_alignment_and_does_not_reuse_one_solver_frame(self) -> None:
        from tempfile import TemporaryDirectory
        from pathlib import Path

        with TemporaryDirectory(prefix="studio-validation-alignment-") as tmp:
            root = Path(tmp)
            gt_path = root / "gt.jsonl"
            trace_path = root / "trace.jsonl"
            _write_jsonl(
                gt_path,
                [
                    {
                        "run_id": "r1",
                        "run_index": 0,
                        "frame_id": 0,
                        "timestamp_ms": 1000,
                        "target_x": 10,
                        "target_y": 20,
                    },
                    {
                        "run_id": "r1",
                        "run_index": 0,
                        "frame_id": 1,
                        "timestamp_ms": 1010,
                        "target_x": 10,
                        "target_y": 20,
                    },
                ],
            )
            _write_jsonl(
                trace_path,
                [
                    {
                        "type": "SOLVER_VISUAL_TRACE",
                        "frame_index": 77,
                        "timestamp_ms": 1005,
                        "payload": {
                            "selected_x": 10,
                            "selected_y": 20,
                            "mouse_enabled": False,
                            "candidate_count": 1,
                        },
                    }
                ],
            )

            result = score_studio_session(
                gt_path,
                trace_path,
                root / "report",
                pass_distance_px=10.0,
                max_alignment_ms=50.0,
            )

            self.assertEqual(result.summary.total_frames, 2)
            self.assertEqual(result.summary.aligned_frames, 1)
            self.assertEqual(result.summary.alignment_missing_frames, 1)
            self.assertEqual(result.summary.alignment_rate, 0.5)
            self.assertEqual(result.summary.aligned_pass_rate, 1.0)
            self.assertEqual(result.runs[0].aligned_frames, 1)
            self.assertEqual(result.runs[0].alignment_rate, 0.5)
            self.assertEqual(result.runs[0].aligned_pass_rate, 1.0)
            report = result.report_path.read_text(encoding="utf-8")
            self.assertIn("alignment_rate: 0.5000", report)
            self.assertIn("aligned_pass_rate: 1.0000", report)

    def test_score_classifies_missing_choice_wrong_candidate_and_path_jump(self) -> None:
        from tempfile import TemporaryDirectory
        from pathlib import Path

        with TemporaryDirectory(prefix="studio-validation-class-") as tmp:
            root = Path(tmp)
            gt_path = root / "gt.jsonl"
            trace_path = root / "trace.jsonl"
            gt_rows = [
                {
                    "run_id": "r1",
                    "run_index": 0,
                    "frame_id": index,
                    "timestamp_ms": 1000 + index * 100,
                    "target_x": float(index),
                    "target_y": 0.0,
                }
                for index in range(5)
            ]
            solver_payloads = [
                {"selected_x": 0.0, "selected_y": 0.0, "candidate_count": 2},
                {"selected_x": 30.0, "selected_y": 0.0, "candidate_count": 2},
                {"selected_x": 200.0, "selected_y": 0.0, "candidate_count": 2},
                {"selected_x": None, "selected_y": None, "candidate_count": 0},
                {"selected_x": None, "selected_y": None, "candidate_count": 3},
            ]
            trace_rows = [
                {
                    "type": "SOLVER_VISUAL_TRACE",
                    "frame_index": index,
                    "timestamp_ms": 1000 + index * 100,
                    "payload": dict(payload, mouse_enabled=False),
                }
                for index, payload in enumerate(solver_payloads)
            ]
            _write_jsonl(gt_path, gt_rows)
            _write_jsonl(trace_path, trace_rows)

            result = score_studio_session(gt_path, trace_path, root / "report", pass_distance_px=10.0)

            self.assertEqual(
                ["", "wrong_candidate", "path_jump", "candidate_missing", "selector_no_choice"],
                [frame.failure_class for frame in result.frames],
            )

    def test_score_reports_phase_failure_stage_and_candidate_judges(self) -> None:
        from tempfile import TemporaryDirectory
        from pathlib import Path

        with TemporaryDirectory(prefix="studio-validation-diagnosis-") as tmp:
            root = Path(tmp)
            gt_path = root / "gt.jsonl"
            trace_path = root / "trace.jsonl"
            _write_jsonl(
                gt_path,
                [
                    {
                        "run_id": "r1",
                        "run_index": 0,
                        "frame_id": 0,
                        "solver_frame_index": 0,
                        "target_x": 10.0,
                        "target_y": 10.0,
                        "target_white": 0.5,
                        "target_motion_started": True,
                        "target_decoy_overlap": True,
                    }
                ],
            )
            ranking = [
                {
                    "candidate_id": "wrong",
                    "center": [80.0, 80.0],
                    "score": 0.8,
                    "distance": 12.0,
                    "total_cost": 18.0,
                    "cost_parts": {"continuity": 12.0, "yolo": 0.0},
                    "judge_shares": {"continuity": 100.0, "yolo": 0.0},
                }
            ]
            _write_jsonl(
                trace_path,
                [
                    {
                        "type": "CANDIDATES",
                        "frame_index": 0,
                        "payload": {
                            "count": 1,
                            "candidates": [
                                {
                                    "candidate_id": "wrong",
                                    "bbox": [75.0, 75.0, 85.0, 85.0],
                                    "center": [80.0, 80.0],
                                    "score": 0.8,
                                    "source": "raw",
                                }
                            ],
                        },
                    },
                    {
                        "type": "EVIDENCE",
                        "frame_index": 0,
                        "payload": {
                            "evidence": [
                                {
                                    "candidate_id": "wrong",
                                    "bg_score": 0.7,
                                    "motion_divergence": 0.2,
                                    "rigid_violation": 0.3,
                                    "phase_similarity": 0.8,
                                    "texture_bg_score": 0.9,
                                    "color_residual": 0.1,
                                    "merge_likelihood": 0.6,
                                }
                            ]
                        },
                    },
                    {
                        "type": "IDENTITY_STATE",
                        "frame_index": 0,
                        "payload": {
                            "candidate_id": "wrong",
                            "reason": "reacquired",
                            "debug": {"ranking": ranking},
                        },
                    },
                    {
                        "type": "TARGET_SELECTION",
                        "frame_index": 0,
                        "payload": {"point": [80.0, 80.0]},
                    },
                    {
                        "type": "SOLVER_VISUAL_TRACE",
                        "frame_index": 0,
                        "payload": {
                            "selected_x": 80.0,
                            "selected_y": 80.0,
                            "candidate_id": "wrong",
                            "reason": "reacquired",
                            "mouse_enabled": False,
                            "candidate_count": 1,
                        },
                    },
                ],
            )

            result = score_studio_session(gt_path, trace_path, root / "report", pass_distance_px=10.0)

            self.assertEqual("fade", result.frames[0].puzzle_phase)
            self.assertEqual("reacquire", result.frames[0].failure_stage)
            self.assertTrue(result.frames[0].target_overlap)
            self.assertEqual(2, len(result.candidate_diagnostics))
            selected = next(row for row in result.candidate_diagnostics if row.role == "selected_identity")
            self.assertEqual("wrong", selected.candidate_id)
            self.assertEqual(18.0, selected.total_cost)
            workbook = load_workbook(result.xlsx_path)
            self.assertIn("candidate_diagnostics", workbook.sheetnames)

    def test_diagnostic_frame_marks_solver_gt_and_candidate_boxes(self) -> None:
        import numpy as np

        frame = np.zeros((120, 160, 3), dtype=np.uint8)
        annotated = _annotate_diagnostic_frame(
            frame,
            frame_id=77,
            puzzle_phase="fade",
            failure_stage="reacquire",
            distance_px=72.8,
            selected_point=(80.0, 60.0),
            gt_point=(100.0, 40.0),
            selected_bbox=(70.0, 50.0, 90.0, 70.0),
            nearest_bbox=(92.0, 32.0, 108.0, 48.0),
        )

        self.assertEqual(frame.shape, annotated.shape)
        self.assertGreater(int(annotated.sum()), 0)

    def test_representative_failure_images_are_limited_to_four_per_run(self) -> None:
        from types import SimpleNamespace

        frames = [
            SimpleNamespace(
                run_id="r1",
                frame_id=index,
                distance_px=float(index * 10),
                failure_stage=("overlap" if index == 3 else "reacquire"),
            )
            for index in range(1, 9)
        ]

        keys = _representative_failure_keys(frames)

        self.assertLessEqual(len(keys), 4)
        self.assertIn(("r1", 1), keys)
        self.assertIn(("r1", 8), keys)
        self.assertIn(("r1", 3), keys)


if __name__ == "__main__":
    unittest.main()
