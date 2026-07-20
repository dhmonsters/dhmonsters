# Studio 가설 보관 실패의 원인을 자동 분류하는 감사기를 검증합니다.
from __future__ import annotations

import json
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from core.puzzle.studio_hypothesis_audit import audit_hypothesis_generation


def _write_jsonl(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8") as stream:
        for row in rows:
            stream.write(json.dumps(row, ensure_ascii=False) + "\n")


class StudioHypothesisAuditTest(unittest.TestCase):
    def test_reports_saturated_duplicate_hypothesis_generation_error(self) -> None:
        with TemporaryDirectory(prefix="studio-hypothesis-audit-") as tmp:
            root = Path(tmp)
            score_path = root / "score.jsonl"
            trace_path = root / "trace.jsonl"
            _write_jsonl(
                score_path,
                [
                    {
                        "run_id": "r1",
                        "frame_id": 7,
                        "solver_frame_index": 10,
                        "target_x": 10.0,
                        "target_y": 10.0,
                        "selected_x": 100.0,
                        "selected_y": 100.0,
                        "passed": False,
                        "puzzle_phase": "reacquire",
                    }
                ],
            )
            _write_jsonl(
                trace_path,
                [
                    {
                        "type": "CANDIDATES",
                        "frame_index": 10,
                        "payload": {
                            "candidates": [
                                {
                                    "candidate_id": "wrong",
                                    "center": [100.0, 100.0],
                                    "bbox": [90.0, 90.0, 110.0, 110.0],
                                    "score": 0.9,
                                    "source": "raw",
                                },
                                {
                                    "candidate_id": "target",
                                    "center": [10.0, 10.0],
                                    "bbox": [5.0, 5.0, 15.0, 15.0],
                                    "score": 0.3,
                                    "source": "raw",
                                },
                            ]
                        },
                    },
                    {
                        "type": "TEMPORAL_SELECTOR",
                        "frame_index": 10,
                        "payload": {
                            "debug": {
                                "kinematic_wide_beam_points": [[100.0, 100.0], [100.0, 100.0]],
                                "kinematic_wide_beam_debug": {
                                    "reason": "beam",
                                    "state_count": 2,
                                },
                            }
                        },
                    },
                ],
            )

            result = audit_hypothesis_generation(
                score_path,
                trace_path,
                root / "audit",
                pass_distance_px=10.0,
                beam_width=2,
            )

            self.assertEqual(result.summary.hypothesis_generation_errors, 1)
            self.assertEqual(result.summary.saturated_errors, 1)
            self.assertEqual(result.summary.duplicate_occupied_errors, 1)
            self.assertEqual(result.frames[0].state_count, 2)
            self.assertEqual(result.frames[0].unique_point_count, 1)
            self.assertEqual(result.frames[0].target_candidate_rank, 2)
            self.assertEqual(result.frames[0].target_candidate_id, "target")
            self.assertTrue(result.report_path.is_file())
            self.assertEqual(
                set(load_workbook(result.xlsx_path).sheetnames),
                {"summary", "frames"},
            )


if __name__ == "__main__":
    unittest.main()
