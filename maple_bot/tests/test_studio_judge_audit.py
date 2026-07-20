# Studio selector 심판 독립성 감사기를 검증한다.
from __future__ import annotations

import json
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from core.puzzle.studio_judge_audit import audit_studio_selector


def _write_jsonl(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8") as stream:
        for row in rows:
            stream.write(json.dumps(row, ensure_ascii=False) + "\n")


class StudioJudgeAuditTest(unittest.TestCase):
    def test_audit_includes_only_failed_frames_with_a_retained_target_hypothesis(self) -> None:
        with TemporaryDirectory(prefix="studio-judge-audit-") as tmp:
            root = Path(tmp)
            gt_path = root / "gt.jsonl"
            trace_path = root / "trace.jsonl"
            _write_jsonl(
                gt_path,
                [
                    {"run_id": "r1", "frame_id": 0, "target_x": 10, "target_y": 10},
                    {"run_id": "r1", "frame_id": 1, "target_x": 20, "target_y": 20},
                ],
            )
            trace_rows: list[dict[str, object]] = []
            frame_rows = (
                (
                    0,
                    [100, 100],
                    [[100, 100], [10, 10]],
                    [
                        {"candidate_id": "wrong", "center": [100, 100], "bbox": [95, 95, 105, 105], "score": 0.8},
                        {"candidate_id": "oracle", "center": [10, 10], "bbox": [5, 5, 15, 15], "score": 0.3},
                    ],
                    [
                        {
                            "candidate_id": "wrong",
                            "bg_score": 0.8,
                            "motion_divergence": 0.2,
                            "rigid_violation": 0.2,
                            "local_rigid_residual": 0.1,
                            "phase_similarity": 0.8,
                            "texture_bg_score": 0.8,
                            "merge_likelihood": 0.2,
                        },
                        {
                            "candidate_id": "oracle",
                            "bg_score": 0.3,
                            "motion_divergence": 0.7,
                            "rigid_violation": 0.7,
                            "local_rigid_residual": 0.9,
                            "phase_similarity": 0.3,
                            "texture_bg_score": 0.2,
                            "merge_likelihood": 0.4,
                        },
                    ],
                ),
                (
                    1,
                    [20, 20],
                    [[20, 20], [80, 80]],
                    [
                        {"candidate_id": "selected", "center": [20, 20], "bbox": [15, 15, 25, 25], "score": 0.8},
                        {"candidate_id": "other", "center": [80, 80], "bbox": [75, 75, 85, 85], "score": 0.6},
                    ],
                    [
                        {"candidate_id": "selected", "motion_divergence": 0.6, "rigid_violation": 0.6, "phase_similarity": 0.4},
                        {"candidate_id": "other", "motion_divergence": 0.3, "rigid_violation": 0.3, "phase_similarity": 0.7},
                    ],
                ),
            )
            for frame_index, selected, hypotheses, candidates, evidence in frame_rows:
                trace_rows.extend(
                    [
                        {"type": "CANDIDATES", "frame_index": frame_index, "payload": {"candidates": candidates}},
                        {"type": "EVIDENCE", "frame_index": frame_index, "payload": {"evidence": evidence}},
                        {
                            "type": "TEMPORAL_SELECTOR",
                            "frame_index": frame_index,
                            "payload": {"debug": {"kinematic_wide_beam_points": hypotheses}},
                        },
                        {"type": "TARGET_SELECTION", "frame_index": frame_index, "payload": {"point": selected}},
                        {
                            "type": "SOLVER_VISUAL_TRACE",
                            "frame_index": frame_index,
                            "payload": {
                                "selected_x": selected[0],
                                "selected_y": selected[1],
                                "mouse_enabled": False,
                                "candidate_count": len(candidates),
                            },
                        },
                    ]
                )
            _write_jsonl(trace_path, trace_rows)

            result = audit_studio_selector(gt_path, trace_path, root / "audit", pass_distance_px=10.0)

            self.assertEqual(result.summary.selector_error_frames, 1)
            self.assertEqual(result.summary.audited_pairs, 1)
            self.assertEqual(result.summary.candidate_samples, 2)
            self.assertAlmostEqual(result.summary.motion_rigid_equal_rate, 1.0)
            self.assertAlmostEqual(result.summary.phase_motion_complement_rate, 1.0)
            self.assertAlmostEqual(result.summary.local_rigid_nonzero_rate, 1.0)
            self.assertAlmostEqual(result.summary.local_rigid_motion_equal_rate, 0.0)
            self.assertEqual(result.pairs[0].selected_candidate_id, "wrong")
            self.assertEqual(result.pairs[0].oracle_candidate_id, "oracle")
            self.assertAlmostEqual(result.pairs[0].motion_delta, 0.5)
            self.assertAlmostEqual(result.pairs[0].local_rigid_delta, 0.8)
            self.assertIn("|local_rigid|0.800000|1.0000|", result.report_path.read_text(encoding="utf-8"))
            self.assertTrue(result.report_path.is_file())
            self.assertEqual(
                set(load_workbook(result.xlsx_path).sheetnames),
                {"summary", "pairs", "correlations"},
            )


if __name__ == "__main__":
    unittest.main()
