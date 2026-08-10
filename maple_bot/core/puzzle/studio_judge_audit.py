# Studio selector 오류에서 심판 신호의 독립성과 판별력을 감사한다.
from __future__ import annotations

import argparse
from dataclasses import asdict, dataclass
from itertools import combinations
from math import hypot, sqrt
from pathlib import Path
from statistics import fmean
from typing import Any, Callable

from openpyxl import Workbook

from .studio_validation import _read_jsonl, _retained_hypothesis_points, score_studio_session


FEATURES = (
    "yolo_score",
    "bg_score",
    "motion_divergence",
    "rigid_violation",
    "local_rigid_residual",
    "phase_similarity",
    "texture_bg_score",
    "merge_likelihood",
)


@dataclass(frozen=True)
class JudgeAuditSummary:
    selector_error_frames: int
    audited_pairs: int
    candidate_samples: int
    motion_rigid_equal_rate: float
    phase_motion_complement_rate: float
    local_rigid_nonzero_rate: float
    local_rigid_motion_equal_rate: float


@dataclass(frozen=True)
class JudgePairRow:
    run_id: str
    frame_id: int
    solver_frame_index: int
    selected_candidate_id: str
    oracle_candidate_id: str
    yolo_delta: float
    bg_delta: float
    motion_delta: float
    rigid_delta: float
    local_rigid_delta: float
    phase_delta: float
    texture_delta: float
    merge_delta: float


@dataclass(frozen=True)
class CorrelationRow:
    left_feature: str
    right_feature: str
    correlation: float
    samples: int


@dataclass(frozen=True)
class JudgeAuditResult:
    summary: JudgeAuditSummary
    pairs: tuple[JudgePairRow, ...]
    correlations: tuple[CorrelationRow, ...]
    report_path: Path
    xlsx_path: Path


def audit_studio_selector(
    gt_jsonl: str | Path,
    trace_jsonl: str | Path,
    output_dir: str | Path,
    *,
    pass_distance_px: float = 24.0,
) -> JudgeAuditResult:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    validation = score_studio_session(
        gt_jsonl,
        trace_jsonl,
        output / "validation",
        pass_distance_px=pass_distance_px,
    )
    trace_rows = _read_jsonl(Path(trace_jsonl))
    events = _events_by_frame(trace_rows)
    pairs: list[JudgePairRow] = []
    samples: list[dict[str, float]] = []
    selector_error_frames = 0

    for frame in validation.frames:
        if (
            frame.passed
            or frame.solver_frame_index is None
            or frame.target_x is None
            or frame.target_y is None
        ):
            continue
        frame_index = frame.solver_frame_index
        temporal_payload = events.get((frame_index, "TEMPORAL_SELECTOR"), {})
        retained_points = _retained_hypothesis_points(temporal_payload)
        oracle_points = [
            point
            for point in retained_points
            if _distance(point, (frame.target_x, frame.target_y)) <= pass_distance_px
        ]
        if not oracle_points:
            continue
        selector_error_frames += 1
        candidate_payload = events.get((frame_index, "CANDIDATES"), {})
        evidence_payload = events.get((frame_index, "EVIDENCE"), {})
        target_payload = events.get((frame_index, "TARGET_SELECTION"), {})
        candidates = [
            row
            for row in candidate_payload.get("candidates", [])
            if isinstance(row, dict)
        ]
        evidence_by_id = {
            str(row.get("candidate_id", "")): row
            for row in evidence_payload.get("evidence", [])
            if isinstance(row, dict)
        }
        for candidate in candidates:
            candidate_id = str(candidate.get("candidate_id", ""))
            evidence = evidence_by_id.get(candidate_id)
            if evidence is not None:
                samples.append(_feature_values(candidate, evidence))

        selected_point = _point(target_payload.get("point"))
        oracle_point = min(
            oracle_points,
            key=lambda point: _distance(point, (frame.target_x, frame.target_y)),
        )
        selected_candidate = _nearest_candidate(candidates, selected_point)
        oracle_candidate = _nearest_candidate(candidates, oracle_point)
        if selected_candidate is None or oracle_candidate is None:
            continue
        selected_id = str(selected_candidate.get("candidate_id", ""))
        oracle_id = str(oracle_candidate.get("candidate_id", ""))
        selected_evidence = evidence_by_id.get(selected_id)
        oracle_evidence = evidence_by_id.get(oracle_id)
        if selected_evidence is None or oracle_evidence is None:
            continue
        selected_values = _feature_values(selected_candidate, selected_evidence)
        oracle_values = _feature_values(oracle_candidate, oracle_evidence)
        pairs.append(
            JudgePairRow(
                run_id=frame.run_id,
                frame_id=frame.frame_id,
                solver_frame_index=frame_index,
                selected_candidate_id=selected_id,
                oracle_candidate_id=oracle_id,
                yolo_delta=oracle_values["yolo_score"] - selected_values["yolo_score"],
                bg_delta=oracle_values["bg_score"] - selected_values["bg_score"],
                motion_delta=oracle_values["motion_divergence"] - selected_values["motion_divergence"],
                rigid_delta=oracle_values["rigid_violation"] - selected_values["rigid_violation"],
                local_rigid_delta=oracle_values["local_rigid_residual"]
                - selected_values["local_rigid_residual"],
                phase_delta=oracle_values["phase_similarity"] - selected_values["phase_similarity"],
                texture_delta=oracle_values["texture_bg_score"] - selected_values["texture_bg_score"],
                merge_delta=oracle_values["merge_likelihood"] - selected_values["merge_likelihood"],
            )
        )

    motion_rigid_rows = [
        row
        for row in samples
        if "motion_divergence" in row and "rigid_violation" in row
    ]
    phase_motion_rows = [
        row
        for row in samples
        if "phase_similarity" in row and "motion_divergence" in row
    ]
    local_rigid_rows = [
        row
        for row in samples
        if "local_rigid_residual" in row and "motion_divergence" in row
    ]
    summary = JudgeAuditSummary(
        selector_error_frames=selector_error_frames,
        audited_pairs=len(pairs),
        candidate_samples=len(samples),
        motion_rigid_equal_rate=_rate(
            motion_rigid_rows,
            lambda row: abs(row["motion_divergence"] - row["rigid_violation"]) <= 1e-9,
        ),
        phase_motion_complement_rate=_rate(
            phase_motion_rows,
            lambda row: abs(row["phase_similarity"] + row["motion_divergence"] - 1.0) <= 1e-6,
        ),
        local_rigid_nonzero_rate=_rate(
            local_rigid_rows,
            lambda row: row["local_rigid_residual"] > 1e-9,
        ),
        local_rigid_motion_equal_rate=_rate(
            local_rigid_rows,
            lambda row: abs(row["local_rigid_residual"] - row["motion_divergence"]) <= 1e-9,
        ),
    )
    correlations = tuple(_correlations(samples))
    report_path = output / "studio_judge_audit.md"
    xlsx_path = output / "studio_judge_audit.xlsx"
    report_path.write_text(_render_report(summary, pairs, correlations), encoding="utf-8")
    _write_xlsx(xlsx_path, summary, pairs, correlations)
    return JudgeAuditResult(
        summary=summary,
        pairs=tuple(pairs),
        correlations=correlations,
        report_path=report_path,
        xlsx_path=xlsx_path,
    )


def _events_by_frame(trace_rows: list[dict[str, Any]]) -> dict[tuple[int, str], dict[str, Any]]:
    events: dict[tuple[int, str], dict[str, Any]] = {}
    for row in trace_rows:
        frame_index = row.get("frame_index")
        payload = row.get("payload")
        if isinstance(frame_index, int) and isinstance(payload, dict):
            events[(frame_index, str(row.get("type", "")))] = payload
    return events


def _feature_values(candidate: dict[str, Any], evidence: dict[str, Any]) -> dict[str, float]:
    return {
        "yolo_score": _float(candidate.get("score")),
        "bg_score": _float(evidence.get("bg_score")),
        "motion_divergence": _float(evidence.get("motion_divergence")),
        "rigid_violation": _float(evidence.get("rigid_violation")),
        "local_rigid_residual": _float(evidence.get("local_rigid_residual")),
        "phase_similarity": _float(evidence.get("phase_similarity")),
        "texture_bg_score": _float(evidence.get("texture_bg_score")),
        "merge_likelihood": _float(evidence.get("merge_likelihood")),
    }


def _point(value: object) -> tuple[float, float] | None:
    if not isinstance(value, (list, tuple)) or len(value) < 2:
        return None
    try:
        return (float(value[0]), float(value[1]))
    except (TypeError, ValueError):
        return None


def _nearest_candidate(
    candidates: list[dict[str, Any]],
    point: tuple[float, float] | None,
) -> dict[str, Any] | None:
    if point is None:
        return None
    rows = [
        (candidate, _point(candidate.get("center")))
        for candidate in candidates
    ]
    valid = [(candidate, center) for candidate, center in rows if center is not None]
    if not valid:
        return None
    return min(valid, key=lambda row: _distance(row[1], point))[0]


def _distance(left: tuple[float, float], right: tuple[float, float]) -> float:
    return hypot(left[0] - right[0], left[1] - right[1])


def _float(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _rate(rows: list[dict[str, float]], predicate: Callable[[dict[str, float]], bool]) -> float:
    return sum(1 for row in rows if predicate(row)) / len(rows) if rows else 0.0


def _correlations(samples: list[dict[str, float]]) -> list[CorrelationRow]:
    rows: list[CorrelationRow] = []
    for left, right in combinations(FEATURES, 2):
        values = [(sample[left], sample[right]) for sample in samples]
        rows.append(
            CorrelationRow(
                left_feature=left,
                right_feature=right,
                correlation=_pearson(values),
                samples=len(values),
            )
        )
    return rows


def _pearson(values: list[tuple[float, float]]) -> float:
    if len(values) < 2:
        return 0.0
    left_mean = fmean(value[0] for value in values)
    right_mean = fmean(value[1] for value in values)
    numerator = sum((left - left_mean) * (right - right_mean) for left, right in values)
    left_scale = sqrt(sum((left - left_mean) ** 2 for left, _right in values))
    right_scale = sqrt(sum((right - right_mean) ** 2 for _left, right in values))
    denominator = left_scale * right_scale
    return numerator / denominator if denominator > 1e-12 else 0.0


def _render_report(
    summary: JudgeAuditSummary,
    pairs: list[JudgePairRow],
    correlations: tuple[CorrelationRow, ...],
) -> str:
    lines = [
        "# Studio Selector Judge Audit",
        "",
        "- GT is used only for post-run audit scoring.",
        f"- selector_error_frames: {summary.selector_error_frames}",
        f"- audited_pairs: {summary.audited_pairs}",
        f"- candidate_samples: {summary.candidate_samples}",
        f"- motion_rigid_equal_rate: {summary.motion_rigid_equal_rate:.4f}",
        f"- phase_motion_complement_rate: {summary.phase_motion_complement_rate:.4f}",
        f"- local_rigid_nonzero_rate: {summary.local_rigid_nonzero_rate:.4f}",
        f"- local_rigid_motion_equal_rate: {summary.local_rigid_motion_equal_rate:.4f}",
        "",
        "## Oracle Pair Signals",
        "",
        "|signal|mean oracle-selected delta|oracle not worse rate|",
        "|---|---:|---:|",
    ]
    signal_rows = (
        ("yolo", [row.yolo_delta for row in pairs], lambda value: value >= 0.0),
        ("background", [row.bg_delta for row in pairs], lambda value: value <= 0.0),
        ("motion", [row.motion_delta for row in pairs], lambda value: value >= 0.0),
        ("rigid", [row.rigid_delta for row in pairs], lambda value: value >= 0.0),
        ("local_rigid", [row.local_rigid_delta for row in pairs], lambda value: value >= 0.0),
        ("phase", [row.phase_delta for row in pairs], lambda value: value <= 0.0),
        ("texture", [row.texture_delta for row in pairs], lambda value: value <= 0.0),
        ("merge", [row.merge_delta for row in pairs], lambda value: value >= 0.0),
    )
    for name, values, favorable in signal_rows:
        mean = fmean(values) if values else 0.0
        rate = sum(1 for value in values if favorable(value)) / len(values) if values else 0.0
        lines.append(f"|{name}|{mean:.6f}|{rate:.4f}|")
    lines.extend(
        [
            "",
            "## Strong Correlations",
            "",
            "|left|right|correlation|samples|",
            "|---|---|---:|---:|",
        ]
    )
    for row in sorted(correlations, key=lambda item: abs(item.correlation), reverse=True)[:12]:
        lines.append(
            f"|{row.left_feature}|{row.right_feature}|{row.correlation:.6f}|{row.samples}|"
        )
    return "\n".join(lines) + "\n"


def _write_xlsx(
    path: Path,
    summary: JudgeAuditSummary,
    pairs: list[JudgePairRow],
    correlations: tuple[CorrelationRow, ...],
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["metric", "value"])
    for key, value in asdict(summary).items():
        summary_sheet.append([key, value])
    _append_rows(workbook.create_sheet("pairs"), pairs, JudgePairRow)
    _append_rows(workbook.create_sheet("correlations"), list(correlations), CorrelationRow)
    workbook.save(path)


def _append_rows(sheet: Any, rows: list[Any], row_type: type[Any]) -> None:
    headers = list(row_type.__dataclass_fields__)
    sheet.append(headers)
    for row in rows:
        values = asdict(row)
        sheet.append([values[header] for header in headers])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("gt_jsonl", type=Path)
    parser.add_argument("trace_jsonl", type=Path)
    parser.add_argument("output_dir", type=Path)
    parser.add_argument("--pass-distance-px", type=float, default=24.0)
    args = parser.parse_args()
    result = audit_studio_selector(
        args.gt_jsonl,
        args.trace_jsonl,
        args.output_dir,
        pass_distance_px=args.pass_distance_px,
    )
    print(result.report_path)
    print(result.xlsx_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
