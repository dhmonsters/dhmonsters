# Studio에서 검출된 정답 후보가 시간축 보관함에서 빠지는 원인을 분석합니다.
from __future__ import annotations

import argparse
from dataclasses import asdict, dataclass
import json
from math import hypot
from pathlib import Path
from statistics import fmean
from typing import Any

from openpyxl import Workbook

from .studio_validation import _read_jsonl, _retained_hypothesis_points


@dataclass(frozen=True)
class HypothesisAuditSummary:
    hypothesis_generation_errors: int
    saturated_errors: int
    duplicate_occupied_errors: int
    mean_state_count: float
    mean_unique_point_count: float
    mean_target_candidate_rank: float


@dataclass(frozen=True)
class HypothesisAuditFrame:
    run_id: str
    frame_id: int
    solver_frame_index: int
    puzzle_phase: str
    state_count: int
    unique_point_count: int
    saturated: bool
    duplicate_occupied: bool
    target_candidate_id: str
    target_candidate_rank: int
    target_candidate_score: float
    target_candidate_source: str
    target_candidate_distance_px: float
    nearest_retained_distance_px: float
    beam_reason: str


@dataclass(frozen=True)
class HypothesisAuditResult:
    summary: HypothesisAuditSummary
    frames: tuple[HypothesisAuditFrame, ...]
    report_path: Path
    xlsx_path: Path


def audit_hypothesis_suite(
    suite_root: str | Path,
    output_dir: str | Path,
    *,
    pass_distance_px: float = 24.0,
    beam_width: int = 16,
) -> dict[str, object]:
    root = Path(suite_root)
    output = Path(output_dir)
    rows: list[dict[str, object]] = []
    for seed_dir in sorted(root.glob("seed_[0-9][0-9]")):
        score_path = next(seed_dir.rglob("score.jsonl"), None)
        trace_path = next(seed_dir.rglob("trace.jsonl"), None)
        if score_path is None or trace_path is None:
            continue
        result = audit_hypothesis_generation(
            score_path,
            trace_path,
            output / seed_dir.name,
            pass_distance_px=pass_distance_px,
            beam_width=beam_width,
        )
        rows.append({"seed": seed_dir.name, **asdict(result.summary)})
    total = {
        key: sum(int(row[key]) for row in rows)
        for key in (
            "hypothesis_generation_errors",
            "saturated_errors",
            "duplicate_occupied_errors",
        )
    }
    output.mkdir(parents=True, exist_ok=True)
    report_path = output / "studio_hypothesis_suite.md"
    xlsx_path = output / "studio_hypothesis_suite.xlsx"
    report_path.write_text(_render_suite_report(rows, total), encoding="utf-8")
    _write_suite_xlsx(xlsx_path, rows, total)
    return {"rows": rows, "total": total, "report_path": report_path, "xlsx_path": xlsx_path}


def audit_hypothesis_generation(
    score_jsonl: str | Path,
    trace_jsonl: str | Path,
    output_dir: str | Path,
    *,
    pass_distance_px: float = 24.0,
    beam_width: int = 16,
) -> HypothesisAuditResult:
    score_rows = _read_jsonl(Path(score_jsonl))
    trace_rows = _read_jsonl(Path(trace_jsonl))
    events = _events_by_frame(trace_rows)
    frames: list[HypothesisAuditFrame] = []

    for score in score_rows:
        frame_index = _optional_int(score.get("solver_frame_index"))
        target = _point((score.get("target_x"), score.get("target_y")))
        if bool(score.get("passed")) or frame_index is None or target is None:
            continue
        candidate_payload = events.get((frame_index, "CANDIDATES"), {})
        temporal_payload = events.get((frame_index, "TEMPORAL_SELECTOR"), {})
        candidates = [
            candidate
            for candidate in candidate_payload.get("candidates", [])
            if isinstance(candidate, dict) and _point(candidate.get("center")) is not None
        ]
        if not candidates:
            continue
        target_candidate = min(
            candidates,
            key=lambda candidate: _distance(_point(candidate.get("center")), target),
        )
        target_distance = _distance(_point(target_candidate.get("center")), target)
        if target_distance > pass_distance_px:
            continue
        retained = _retained_hypothesis_points(temporal_payload)
        if any(_distance(point, target) <= pass_distance_px for point in retained):
            continue

        debug = temporal_payload.get("debug", {})
        wide_debug = debug.get("kinematic_wide_beam_debug", {}) if isinstance(debug, dict) else {}
        state_count = _optional_int(wide_debug.get("state_count")) or len(retained)
        unique_point_count = len({_point_key(point) for point in retained})
        frames.append(
            HypothesisAuditFrame(
                run_id=str(score.get("run_id", "")),
                frame_id=_optional_int(score.get("frame_id")) or 0,
                solver_frame_index=frame_index,
                puzzle_phase=str(score.get("puzzle_phase", "")),
                state_count=state_count,
                unique_point_count=unique_point_count,
                saturated=state_count >= max(1, int(beam_width)),
                duplicate_occupied=unique_point_count < state_count,
                target_candidate_id=str(target_candidate.get("candidate_id", "")),
                target_candidate_rank=candidates.index(target_candidate) + 1,
                target_candidate_score=_float(target_candidate.get("score")),
                target_candidate_source=str(target_candidate.get("source", "")),
                target_candidate_distance_px=target_distance,
                nearest_retained_distance_px=min(
                    (_distance(point, target) for point in retained),
                    default=float("inf"),
                ),
                beam_reason=str(wide_debug.get("reason", "")),
            )
        )

    summary = HypothesisAuditSummary(
        hypothesis_generation_errors=len(frames),
        saturated_errors=sum(frame.saturated for frame in frames),
        duplicate_occupied_errors=sum(frame.duplicate_occupied for frame in frames),
        mean_state_count=_mean(frame.state_count for frame in frames),
        mean_unique_point_count=_mean(frame.unique_point_count for frame in frames),
        mean_target_candidate_rank=_mean(frame.target_candidate_rank for frame in frames),
    )
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    report_path = output / "studio_hypothesis_audit.md"
    xlsx_path = output / "studio_hypothesis_audit.xlsx"
    report_path.write_text(_render_report(summary, frames), encoding="utf-8")
    _write_xlsx(xlsx_path, summary, frames)
    return HypothesisAuditResult(summary, tuple(frames), report_path, xlsx_path)


def _events_by_frame(rows: list[dict[str, Any]]) -> dict[tuple[int, str], dict[str, Any]]:
    events: dict[tuple[int, str], dict[str, Any]] = {}
    for row in rows:
        frame_index = _optional_int(row.get("frame_index"))
        payload = row.get("payload")
        if frame_index is not None and isinstance(payload, dict):
            events[(frame_index, str(row.get("type", "")))] = payload
    return events


def _render_report(
    summary: HypothesisAuditSummary,
    frames: list[HypothesisAuditFrame],
) -> str:
    lines = [
        "# Studio 가설 생성 오류 감사",
        "",
        f"- 가설 생성 오류: {summary.hypothesis_generation_errors}",
        f"- 보관함 포화 오류: {summary.saturated_errors}",
        f"- 중복 상태 점유 오류: {summary.duplicate_occupied_errors}",
        f"- 평균 상태 수: {summary.mean_state_count:.2f}",
        f"- 평균 서로 다른 위치 수: {summary.mean_unique_point_count:.2f}",
        f"- 평균 정답 후보 순위: {summary.mean_target_candidate_rank:.2f}",
        "",
        "|run|frame|phase|states|unique|rank|candidate|score|source|nearest retained|",
        "|---|---:|---|---:|---:|---:|---|---:|---|---:|",
    ]
    for frame in frames:
        lines.append(
            f"|{frame.run_id}|{frame.solver_frame_index}|{frame.puzzle_phase}|"
            f"{frame.state_count}|{frame.unique_point_count}|{frame.target_candidate_rank}|"
            f"{frame.target_candidate_id}|{frame.target_candidate_score:.3f}|"
            f"{frame.target_candidate_source}|{frame.nearest_retained_distance_px:.2f}|"
        )
    return "\n".join(lines) + "\n"


def _write_xlsx(
    path: Path,
    summary: HypothesisAuditSummary,
    frames: list[HypothesisAuditFrame],
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["metric", "value"])
    for key, value in asdict(summary).items():
        summary_sheet.append([key, value])
    frame_sheet = workbook.create_sheet("frames")
    fields = list(HypothesisAuditFrame.__dataclass_fields__)
    frame_sheet.append(fields)
    for frame in frames:
        row = asdict(frame)
        frame_sheet.append([row[field] for field in fields])
    workbook.save(path)


def _render_suite_report(rows: list[dict[str, object]], total: dict[str, int]) -> str:
    lines = [
        "# Studio 가설 생성 오류 전체 감사",
        "",
        f"- 가설 생성 오류: {total['hypothesis_generation_errors']}",
        f"- 보관함 포화 오류: {total['saturated_errors']}",
        f"- 중복 상태 점유 오류: {total['duplicate_occupied_errors']}",
        "",
        "|seed|errors|saturated|duplicates|mean states|mean unique|mean rank|",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for row in rows:
        lines.append(
            f"|{row['seed']}|{row['hypothesis_generation_errors']}|"
            f"{row['saturated_errors']}|{row['duplicate_occupied_errors']}|"
            f"{float(row['mean_state_count']):.2f}|"
            f"{float(row['mean_unique_point_count']):.2f}|"
            f"{float(row['mean_target_candidate_rank']):.2f}|"
        )
    return "\n".join(lines) + "\n"


def _write_suite_xlsx(
    path: Path,
    rows: list[dict[str, object]],
    total: dict[str, int],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "seeds"
    fields = list(rows[0]) if rows else ["seed"]
    sheet.append(fields)
    for row in rows:
        sheet.append([row.get(field) for field in fields])
    total_sheet = workbook.create_sheet("total")
    total_sheet.append(["metric", "value"])
    for key, value in total.items():
        total_sheet.append([key, value])
    workbook.save(path)


def _point(value: object) -> tuple[float, float] | None:
    if not isinstance(value, (list, tuple)) or len(value) < 2:
        return None
    try:
        return (float(value[0]), float(value[1]))
    except (TypeError, ValueError):
        return None


def _distance(left: tuple[float, float] | None, right: tuple[float, float]) -> float:
    if left is None:
        return float("inf")
    return hypot(left[0] - right[0], left[1] - right[1])


def _point_key(point: tuple[float, float]) -> tuple[float, float]:
    return (round(point[0], 4), round(point[1], 4))


def _optional_int(value: object) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _float(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _mean(values: object) -> float:
    rows = list(values)
    return fmean(rows) if rows else 0.0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Studio 가설 생성 오류를 감사합니다.")
    parser.add_argument("--suite-root", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--pass-distance-px", type=float, default=24.0)
    parser.add_argument("--beam-width", type=int, default=16)
    args = parser.parse_args(argv)
    result = audit_hypothesis_suite(
        args.suite_root,
        args.output_dir,
        pass_distance_px=args.pass_distance_px,
        beam_width=args.beam_width,
    )
    print(json.dumps({"rows": result["rows"], "total": result["total"]}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
