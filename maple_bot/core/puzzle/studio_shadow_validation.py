# 동일 Studio 프레임에서 기본 경로와 shadow 경로를 공정하게 사후 채점합니다.
from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from math import hypot
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook


@dataclass(frozen=True)
class ShadowAbFrame:
    run_id: str
    run_index: int
    frame_id: int
    solver_frame_index: int
    gt_x: float
    gt_y: float
    base_x: float | None
    base_y: float | None
    shadow_x: float | None
    shadow_y: float | None
    base_distance_px: float | None
    shadow_distance_px: float | None
    base_passed: bool
    shadow_passed: bool
    outcome: str
    gate_available: bool
    gate_selected: bool
    gate_reason: str


@dataclass(frozen=True)
class ShadowAbRunSummary:
    run_id: str
    run_index: int
    total_frames: int
    base_passed_frames: int
    shadow_passed_frames: int
    improved_frames: int
    regressed_frames: int
    delta_frames: int


@dataclass(frozen=True)
class ShadowAbSummary:
    total_frames: int
    aligned_frames: int
    gate_available_frames: int
    base_passed_frames: int
    shadow_passed_frames: int
    improved_frames: int
    regressed_frames: int
    delta_frames: int
    base_pass_rate: float
    shadow_pass_rate: float
    regressed_runs: int
    no_run_regression: bool


@dataclass(frozen=True)
class ShadowAbResult:
    summary: ShadowAbSummary
    runs: list[ShadowAbRunSummary]
    frames: list[ShadowAbFrame]
    score_jsonl: Path
    xlsx_path: Path
    report_path: Path


def score_kinematic_shadow_ab(
    gt_jsonl: str | Path,
    trace_jsonl: str | Path,
    output_dir: str | Path,
    *,
    pass_distance_px: float = 24.0,
) -> ShadowAbResult:
    if pass_distance_px < 0:
        raise ValueError("pass_distance_px must be non-negative")
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    target_rows = _target_rows_by_frame(Path(trace_jsonl))
    frames = [
        _score_frame(row, target_rows.get(int(row["solver_frame_index"])), pass_distance_px)
        for row in _read_jsonl(Path(gt_jsonl))
        if row.get("solver_frame_index") is not None
    ]
    runs = _summarize_runs(frames)
    summary = _summarize(frames, runs)
    score_jsonl = output / "studio_shadow_ab_score.jsonl"
    xlsx_path = output / "studio_shadow_ab.xlsx"
    report_path = output / "studio_shadow_ab.md"
    _write_jsonl(score_jsonl, frames)
    _write_xlsx(xlsx_path, summary, runs, frames)
    report_path.write_text(_render_report(summary, runs), encoding="utf-8")
    return ShadowAbResult(summary, runs, frames, score_jsonl, xlsx_path, report_path)


def _score_frame(
    gt: dict[str, object],
    target: dict[str, object] | None,
    pass_distance_px: float,
) -> ShadowAbFrame:
    target = target or {}
    texture_gate = target.get("kinematic_texture_gate")
    texture_gate = texture_gate if isinstance(texture_gate, dict) else {}
    beam_gate = target.get("kinematic_beam_gate")
    beam_gate = beam_gate if isinstance(beam_gate, dict) else {}
    gate = beam_gate if bool(beam_gate.get("available", False)) else texture_gate
    actual_point = _point(target.get("point"))
    base_point = _point(gate.get("base_point")) or actual_point
    shadow_point = _point(gate.get("selected_point")) or actual_point
    gt_point = (float(gt["target_x"]), float(gt["target_y"]))
    base_distance = _distance(base_point, gt_point)
    shadow_distance = _distance(shadow_point, gt_point)
    base_passed = base_distance is not None and base_distance <= pass_distance_px
    shadow_passed = shadow_distance is not None and shadow_distance <= pass_distance_px
    if shadow_passed and not base_passed:
        outcome = "improved"
    elif base_passed and not shadow_passed:
        outcome = "regressed"
    elif base_passed:
        outcome = "both_passed"
    elif base_distance is None or shadow_distance is None:
        outcome = "missing"
    else:
        outcome = "both_failed"
    return ShadowAbFrame(
        run_id=str(gt.get("run_id") or ""),
        run_index=int(gt.get("run_index", 0)),
        frame_id=int(gt.get("frame_id", 0)),
        solver_frame_index=int(gt["solver_frame_index"]),
        gt_x=gt_point[0],
        gt_y=gt_point[1],
        base_x=base_point[0] if base_point is not None else None,
        base_y=base_point[1] if base_point is not None else None,
        shadow_x=shadow_point[0] if shadow_point is not None else None,
        shadow_y=shadow_point[1] if shadow_point is not None else None,
        base_distance_px=base_distance,
        shadow_distance_px=shadow_distance,
        base_passed=base_passed,
        shadow_passed=shadow_passed,
        outcome=outcome,
        gate_available=bool(gate.get("available", False)),
        gate_selected=bool(gate.get("selected", False)),
        gate_reason=str(gate.get("reason") or ""),
    )


def _summarize_runs(frames: list[ShadowAbFrame]) -> list[ShadowAbRunSummary]:
    grouped: dict[tuple[int, str], list[ShadowAbFrame]] = {}
    for frame in frames:
        grouped.setdefault((frame.run_index, frame.run_id), []).append(frame)
    return [
        ShadowAbRunSummary(
            run_id=run_id,
            run_index=run_index,
            total_frames=len(rows),
            base_passed_frames=sum(row.base_passed for row in rows),
            shadow_passed_frames=sum(row.shadow_passed for row in rows),
            improved_frames=sum(row.outcome == "improved" for row in rows),
            regressed_frames=sum(row.outcome == "regressed" for row in rows),
            delta_frames=sum(row.shadow_passed for row in rows) - sum(row.base_passed for row in rows),
        )
        for (run_index, run_id), rows in sorted(grouped.items())
    ]


def _summarize(frames: list[ShadowAbFrame], runs: list[ShadowAbRunSummary]) -> ShadowAbSummary:
    total = len(frames)
    base_passed = sum(frame.base_passed for frame in frames)
    shadow_passed = sum(frame.shadow_passed for frame in frames)
    regressed_runs = sum(run.delta_frames < 0 for run in runs)
    return ShadowAbSummary(
        total_frames=total,
        aligned_frames=sum(frame.base_distance_px is not None and frame.shadow_distance_px is not None for frame in frames),
        gate_available_frames=sum(frame.gate_available for frame in frames),
        base_passed_frames=base_passed,
        shadow_passed_frames=shadow_passed,
        improved_frames=sum(frame.outcome == "improved" for frame in frames),
        regressed_frames=sum(frame.outcome == "regressed" for frame in frames),
        delta_frames=shadow_passed - base_passed,
        base_pass_rate=base_passed / total if total else 0.0,
        shadow_pass_rate=shadow_passed / total if total else 0.0,
        regressed_runs=regressed_runs,
        no_run_regression=regressed_runs == 0,
    )


def _target_rows_by_frame(path: Path) -> dict[int, dict[str, object]]:
    rows: dict[int, dict[str, object]] = {}
    for event in _read_jsonl(path):
        if event.get("type") != "TARGET_SELECTION" or event.get("frame_index") is None:
            continue
        payload = event.get("payload")
        rows[int(event["frame_index"])] = payload if isinstance(payload, dict) else {}
    return rows


def _point(value: object) -> tuple[float, float] | None:
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes)) or len(value) < 2:
        return None
    try:
        return (float(value[0]), float(value[1]))
    except (TypeError, ValueError):
        return None


def _distance(left: tuple[float, float] | None, right: tuple[float, float]) -> float | None:
    if left is None:
        return None
    return hypot(left[0] - right[0], left[1] - right[1])


def _read_jsonl(path: Path):
    with path.open("r", encoding="utf-8") as stream:
        for line in stream:
            if line.strip():
                yield json.loads(line)


def _write_jsonl(path: Path, frames: list[ShadowAbFrame]) -> None:
    with path.open("w", encoding="utf-8") as stream:
        for frame in frames:
            stream.write(json.dumps(asdict(frame), ensure_ascii=False, separators=(",", ":")))
            stream.write("\n")


def _write_xlsx(
    path: Path,
    summary: ShadowAbSummary,
    runs: list[ShadowAbRunSummary],
    frames: list[ShadowAbFrame],
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["metric", "value"])
    for name, value in asdict(summary).items():
        summary_sheet.append([name, value])
    run_sheet = workbook.create_sheet("Runs")
    run_headers = list(asdict(runs[0]).keys()) if runs else []
    if run_headers:
        run_sheet.append(run_headers)
        for run in runs:
            run_sheet.append(list(asdict(run).values()))
    frame_sheet = workbook.create_sheet("Frames")
    frame_headers = list(asdict(frames[0]).keys()) if frames else []
    if frame_headers:
        frame_sheet.append(frame_headers)
        for frame in frames:
            frame_sheet.append(list(asdict(frame).values()))
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def _render_report(summary: ShadowAbSummary, runs: list[ShadowAbRunSummary]) -> str:
    lines = [
        "# Studio 동일 입력 Shadow A/B",
        "",
        f"- total_frames: {summary.total_frames}",
        f"- aligned_frames: {summary.aligned_frames}",
        f"- gate_available_frames: {summary.gate_available_frames}",
        f"- base_passed_frames: {summary.base_passed_frames}",
        f"- shadow_passed_frames: {summary.shadow_passed_frames}",
        f"- delta_frames: {summary.delta_frames:+d}",
        f"- improved_frames: {summary.improved_frames}",
        f"- regressed_frames: {summary.regressed_frames}",
        f"- regressed_runs: {summary.regressed_runs}",
        f"- no_run_regression: {summary.no_run_regression}",
        "",
        "## Runs",
        "",
    ]
    lines.extend(
        f"- {run.run_id}: base={run.base_passed_frames}, shadow={run.shadow_passed_frames}, delta={run.delta_frames:+d}"
        for run in runs
    )
    return "\n".join(lines) + "\n"
